# processor.py — обновлённая версия с логикой "Большой файл = до color_column включительно"

import time
import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string
from copy import copy

def get_cell_color(cell):
    fill = cell.fill
    if not fill or fill.fgColor is None:
        return None
    fg = fill.fgColor
    if fg.type == 'rgb':
        rgb = fg.rgb
        return rgb[2:].upper() if rgb and rgb.startswith('FF') else rgb.upper() if rgb else None
    elif fg.type == 'theme':
        theme = fg.theme or 0
        tint = round(fg.tint or 0.0, 6)
        return f"THEME_{theme}_{tint}"
    elif fg.type == 'indexed':
        return f"INDEXED_{fg.indexed}"
    return None

def expand_column_range(col_range):
    if not col_range:
        return []
    if ':' not in col_range:
        return [col_range]
    start, end = col_range.split(':')
    start_idx = column_index_from_string(start)
    end_idx = column_index_from_string(end)
    return [get_column_letter(i) for i in range(start_idx, end_idx + 1)]

def process_excel(CONFIG, log_callback=None):
    """
    Основная функция обработки. Принимает CONFIG и опциональный callback для логов.
    """
    def log(msg):
        if log_callback:
            log_callback(msg)
        else:
            print(msg)

    try:
        start = time.perf_counter()

        if CONFIG['output_file'] is None:
            p = Path(CONFIG['input_file'])
            CONFIG['output_file'] = str(p.parent / (p.stem + '_обработанный' + p.suffix))

        log(f"✅ Файл будет прочитан: {CONFIG['input_file']}")
        log(f"💾 Результат сохранится: {CONFIG['output_file']}")

        if os.path.exists(CONFIG['output_file']):
            try:
                os.rename(CONFIG['output_file'], CONFIG['output_file'])
            except PermissionError:
                raise PermissionError(f"Файл открыт в Excel: {CONFIG['output_file']}. Закройте его.")

        wb = load_workbook(CONFIG['input_file'], data_only=True)
        temp_wb = load_workbook(CONFIG['input_file'])
        log(f"✅ Книга загружена. Листы: {wb.sheetnames}")

        elapsed = time.perf_counter() - start
        log(f"⏱️  Время загрузки книги: {elapsed:.3f} сек")

        for sheet_name in (CONFIG['sheet_names'] or wb.sheetnames):
            log(f"\n{'='*60}")
            log(f"📋 ОБРАБОТКА ЛИСТА: '{sheet_name}'")
            log(f"{'='*60}")
            start1 = time.perf_counter()

            ws = wb[sheet_name]
            temp_ws = temp_wb[sheet_name]

            last_row = None
            data_cols = set()

            # --- ОПРЕДЕЛЕНИЕ ДИАПАЗОНА ДАННЫХ ---
            scan_row = CONFIG.get('scan_columns_by_row')

            if scan_row is not None:
                # ✅ НОВАЯ ЛОГИКА: Берём все столбцы от A до color_column включительно
                log(f"🔍 Режим 'Большой файл': сканируем столбцы до '{CONFIG['color_column']}' включительно...")
                color_col_idx = column_index_from_string(CONFIG['color_column'])
                for col_idx in range(1, color_col_idx + 1):  # от A до color_column
                    data_cols.add(col_idx)
                # Находим последнюю строку только в этих столбцах
                for row in range(CONFIG['min_row'], ws.max_row + 1):
                    for col in data_cols:
                        cell = ws.cell(row=row, column=col)
                        if cell.value not in [None, ""]:
                            last_row = row if last_row is None else max(last_row, row)
            else:
                # 📊 Старая логика: сканируем все столбцы по всем строкам
                log("🔍 Сканирование всех столбцов по всем строкам...")
                for row in range(CONFIG['min_row'], ws.max_row + 1):
                    for col in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row, column=col)
                        if cell.value not in [None, ""]:
                            last_row = row if last_row is None else max(last_row, row)
                            data_cols.add(col)

            if last_row is None:
                log("⚠️  Лист пуст — пропускаем.")
                continue

            used_cols = set(data_cols)
            if CONFIG['hierarchy_column'] is not None:
                h_col_idx = column_index_from_string(CONFIG['hierarchy_column'])
                used_cols.add(h_col_idx)
            used_cols = sorted(used_cols)
            used_cols_letters = {get_column_letter(col) for col in used_cols}

            log(f"📏 Диапазон: строки {CONFIG['min_row']}–{last_row}, столбцы: {get_column_letter(used_cols[0])}–{get_column_letter(used_cols[-1])}")

            levels = {}
            if (CONFIG['stages']['hierarchy'] or CONFIG['stages']['grouping']) and CONFIG['hierarchy_column'] is not None:
                log("🔍 Определение уровней по цвету...")

                seen_colors = []
                white_like = [None, 'FFFFFFFF', '00000000']

                for row in range(CONFIG['min_row'], last_row + 1):
                    cell = temp_ws.cell(row=row, column=column_index_from_string(CONFIG['color_column']))
                    color = get_cell_color(cell)
                    if color not in white_like and color not in seen_colors:
                        seen_colors.append(color)

                color_to_level = {color: i + 1 for i, color in enumerate(seen_colors or ['DUMMY'])}
                last_level = len(seen_colors) + 1 if seen_colors else 2
                for white in white_like:
                    color_to_level[white] = last_level

                for row in range(CONFIG['min_row'], last_row + 1):
                    cell = temp_ws.cell(row=row, column=column_index_from_string(CONFIG['color_column']))
                    color = get_cell_color(cell)
                    levels[row] = color_to_level.get(color, last_level)

            if CONFIG['stages']['hierarchy'] and CONFIG['hierarchy_column'] is not None:
                counter = [0] * 10
                h_col_idx = column_index_from_string(CONFIG['hierarchy_column'])
                for row in range(CONFIG['min_row'], last_row + 1):
                    level = levels[row]
                    for i in range(level + 1, 10):
                        counter[i] = 0
                    counter[level] += 1
                    num = '.'.join(str(counter[i]) for i in range(1, level + 1) if counter[i] > 0)
                    cell = ws.cell(row=row, column=h_col_idx)
                    cell.value = num
                    cell.data_type = 's'
                log("✅ Иерархическая нумерация применена")

            if CONFIG['stages']['hierarchy_colors'] and CONFIG['hierarchy_column'] is not None:
                for row in range(CONFIG['min_row'], last_row + 1):
                    temp_color_cell = temp_ws.cell(row=row, column=column_index_from_string(CONFIG['color_column']))
                    has_color = get_cell_color(temp_color_cell) not in [None, 'FFFFFFFF', '00000000']
                    level = levels.get(row, 9)
                    cell = ws.cell(row=row, column=column_index_from_string(CONFIG['hierarchy_column']))
                    if has_color and temp_color_cell.fill:
                        cell.fill = copy(temp_color_cell.fill)
                log("✅ В нумерацию добавлен цвет из оригинального столбца")

            if CONFIG['stages']['grouping'] and CONFIG['hierarchy_column'] is not None:
                for row in range(CONFIG['min_row'], last_row + 1):
                    ws.row_dimensions[row].outlineLevel = 0
                    ws.row_dimensions[row].hidden = False
                    ws.row_dimensions[row].collapsed = False

                for current_level in range(1, 8):
                    i = 0
                    while i < last_row - CONFIG['min_row'] + 1:
                        row = CONFIG['min_row'] + i
                        level = levels.get(row, 9)
                        if level == current_level:
                            end_row = last_row
                            for j in range(i + 1, last_row - CONFIG['min_row'] + 1):
                                next_row = CONFIG['min_row'] + j
                                next_level = levels.get(next_row, 9)
                                if next_level <= current_level:
                                    end_row = next_row - 1
                                    break
                            if row < end_row:
                                ws.row_dimensions[row].outlineLevel = current_level - 1
                                ws.row_dimensions[row].collapsed = True
                                for r in range(row + 1, end_row + 1):
                                    nested_level = levels.get(r, 9)
                                    if nested_level > current_level:
                                        ws.row_dimensions[r].outlineLevel = current_level
                        i += 1

                if hasattr(ws, 'sheet_properties') and hasattr(ws.sheet_properties, 'outlinePr'):
                    ws.sheet_properties.outlinePr.summaryBelow = True
                    ws.sheet_properties.outlinePr.summaryRight = True
                    ws.sheet_properties.outlinePr.showOutlineSymbols = True

                log("✅ Группировка применена")

            if CONFIG['stages']['wrap_text'] and CONFIG['wrap_text_columns']:
                log("🔁 Применение переноса текста...")
                wrap_cols = [
                    col for col in CONFIG['wrap_text_columns']
                    if col in used_cols_letters
                ]
                for row in range(CONFIG['min_row'], last_row + 1):
                    if ws.row_dimensions[row].height is not None:
                        ws.row_dimensions[row].height = None
                    for col_letter in wrap_cols:
                        col_idx = column_index_from_string(col_letter)
                        cell = ws.cell(row=row, column=col_idx)
                        h_align = cell.alignment.horizontal if cell.alignment and cell.alignment.horizontal else 'left'
                        v_align = cell.alignment.vertical if cell.alignment and cell.alignment.vertical else 'bottom'
                        cell.alignment = Alignment(
                            horizontal=h_align,
                            vertical=v_align,
                            wrap_text=True
                        )
                log(f"✅ Перенос текста: {', '.join(wrap_cols)}")

            if CONFIG['stages']['alignment'] and CONFIG['alignment_rules']:
                log("📏 Применение выравнивания...")
                applied_cols = set()
                for rule in CONFIG['alignment_rules']:
                    if len(rule) != 3:
                        continue
                    col_range, vertical, horizontal = rule
                    cols = expand_column_range(col_range)
                    for col_letter in cols:
                        if col_letter in used_cols_letters:
                            col_idx = column_index_from_string(col_letter)
                            for row in range(CONFIG['min_row'], last_row + 1):
                                cell = ws.cell(row=row, column=col_idx)
                                wrap = cell.alignment.wrap_text if cell.alignment else False
                                cell.alignment = Alignment(
                                    vertical=vertical,
                                    horizontal=horizontal,
                                    wrap_text=wrap
                                )
                            applied_cols.add(col_letter)
                log(f"✅ Выравнивание: {', '.join(sorted(applied_cols))}")

            if CONFIG['stages']['formatting']:
                log("🎨 Применение форматирования...")

                font_name = CONFIG['font'].get('name', 'Times New Roman')
                font_size = CONFIG['font'].get('size', 14)
                font_bold = CONFIG['font'].get('bold', False)
                font_italic = CONFIG['font'].get('italic', False)
                font_underline = 'single' if CONFIG['font'].get('underline', False) else None

                base_font = Font(
                    name=font_name,
                    size=font_size,
                    bold=font_bold,
                    italic=font_italic,
                    underline=font_underline
                )

                border_style = CONFIG.get('border_style', 'thin')
                border = Border(
                    left=Side(style=border_style),
                    right=Side(style=border_style),
                    top=Side(style=border_style),
                    bottom=Side(style=border_style)
                )

                for row in range(CONFIG['min_row'], last_row + 1):
                    temp_color_cell = temp_ws.cell(row=row, column=column_index_from_string(CONFIG['color_column']))
                    has_color = get_cell_color(temp_color_cell) not in [None, 'FFFFFFFF', '00000000']
                    level = levels.get(row, 9)

                    for col in used_cols:
                        cell = ws.cell(row=row, column=col)

                        is_bold_level = level in CONFIG.get('bold_levels', [1, 2])
                        current_font = Font(
                            name=font_name,
                            size=font_size,
                            bold=font_bold or is_bold_level,
                            italic=font_italic,
                            underline=font_underline
                        )
                        cell.font = current_font

                        cell.border = border

                        if hasattr(CONFIG, 'fill_color') and CONFIG['fill_color']:
                            color = CONFIG['fill_color'].replace("#", "") if CONFIG['fill_color'].startswith("#") else CONFIG['fill_color']
                            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

                        if hasattr(CONFIG, 'text_color') and CONFIG['text_color']:
                            text_color = CONFIG['text_color'].replace("#", "") if CONFIG['text_color'].startswith("#") else CONFIG['text_color']
                            cell.font = Font(
                                name=font_name,
                                size=font_size,
                                bold=font_bold or is_bold_level,
                                italic=font_italic,
                                underline=font_underline,
                                color=text_color
                            )

                        if has_color and temp_color_cell.fill and CONFIG['stages']['hierarchy_colors']:
                            cell.fill = copy(temp_color_cell.fill)

                log("✅ Форматирование применено")

            if CONFIG['stages']['number_formats'] and CONFIG['column_formats']:
                log("🔢 Применение числовых форматов...")
                for col_range, num_format in CONFIG['column_formats'].items():
                    cols = expand_column_range(col_range)
                    for col_letter in cols:
                        if col_letter in used_cols_letters:
                            col_idx = column_index_from_string(col_letter)
                            for row in range(CONFIG['min_row'], last_row + 1):
                                cell = ws.cell(row=row, column=col_idx)
                                if cell.value is not None:
                                    try:
                                        cell.number_format = num_format
                                    except Exception as e:
                                        log(f"⚠️ Ошибка формата {num_format} в {col_letter}{row}: {e}")
                log("✅ Числовые форматы применены")

            log(f"✅ Лист '{sheet_name}' полностью обработан")

        wb.save(CONFIG['output_file'])
        log(f"\n🎉 УСПЕШНО: файл сохранён!")
        log(f"📁 {CONFIG['output_file']}")

        try:
            test_wb = load_workbook(CONFIG['output_file'], read_only=True)
            log("✅ Тестовый запуск успешен — файл корректен.")
            test_wb.close()
        except Exception as e:
            log(f"⚠️ Ошибка при тестовом открытии: {e}")

        return True, "Обработка завершена успешно."

    except Exception as e:
        error_msg = f"❌ Ошибка: {str(e)}"
        log(error_msg)
        return False, error_msg