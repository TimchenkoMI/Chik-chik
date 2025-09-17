# main.py — Chik-chik — ФИНАЛЬНАЯ ВЕРСИЯ v6.0

import sys
import os
import json
import psutil
import time
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QLineEdit, QFileDialog, QTextEdit, QGroupBox,
    QCheckBox, QSpinBox, QScrollArea, QGridLayout, QComboBox,
    QFontComboBox, QToolButton, QStyle,
    QTabWidget, QTableWidget, QTableWidgetItem, QListWidget, QListWidgetItem,
    QHeaderView, QMessageBox, QFrame, QStyleFactory
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QSize, QSettings, QFileInfo
from PyQt5.QtGui import QFont, QColor, QIcon, QPalette
import qdarkstyle
from processor import process_excel

SETTINGS_FILE = "chikchik_settings.json"


# ======================
# ВСПОМОГАТЕЛЬНЫЕ КЛАССЫ
# ======================

class PreviewLabel(QLabel):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setText("Образец: 123 ABC xyz")
        self.setAlignment(Qt.AlignCenter)
        self.setFrameStyle(QFrame.StyledPanel)
        self.setMinimumHeight(40)
        self.update_preview()

    def update_preview(self, font_name="Times New Roman", font_size=14, bold=False, italic=False, underline=False, text_color=None, bg_color=None):
        font = QFont(font_name, font_size)
        font.setBold(bold)
        font.setItalic(italic)
        font.setUnderline(underline)
        self.setFont(font)

        style = "padding: 5px; border: 1px solid #555;"
        if bg_color:
            style += f" background-color: {bg_color};"
        if text_color:
            style += f" color: {text_color};"
        else:
            style += " color: white;"
        self.setStyleSheet(style)


class NumberFormatPreview(QLabel):
    def __init__(self, parent=None, sample_value=12345.67):
        super().__init__(parent)
        self.sample_value = sample_value
        self.setAlignment(Qt.AlignCenter)
        self.setFrameStyle(QFrame.StyledPanel)
        self.setMinimumHeight(30)
        self.setText("Пример: 12,345.67")
        self.update_preview("#,##0.00")

    def update_preview(self, number_format):
        try:
            fmt = number_format.replace("\\", "").replace('"', '').replace("'", "")

            if "#,##0.00" in fmt:
                formatted = f"{self.sample_value:,.2f}"
            elif "#,##0" in fmt:
                formatted = f"{int(self.sample_value):,}"
            elif "0.00%" in fmt:
                formatted = f"{self.sample_value:.2%}"
            elif "0.00E+00" in fmt:
                formatted = f"{self.sample_value:.2e}"
            elif "@" in fmt:
                formatted = str(self.sample_value)
            elif "0.00" in fmt:
                decimals = fmt.count("0") - (fmt.find(".") + 1) if "." in fmt else 0
                formatted = f"{self.sample_value:.{decimals}f}"
            elif "0" in fmt and "." not in fmt:
                formatted = str(int(self.sample_value))
            else:
                formatted = str(self.sample_value)

            if "₽" in fmt:
                formatted = "₽" + formatted
            if "руб" in fmt:
                formatted = formatted + " руб."
            if "$" in fmt:
                formatted = "$" + formatted

            self.setText(f"→ {formatted}")
            self.setStyleSheet("padding: 3px; background-color: #2d2d2d; color: #aaff77; border: 1px solid #55aa55;")

        except Exception:
            self.setText("❌ Неверный формат")
            self.setStyleSheet("padding: 3px; background-color: #3d2d2d; color: #ff7777; border: 1px solid #aa5555;")


# ======================
# КОНФИГ
# ======================

class Config:
    def __init__(self):
        self.input_file = ""
        self.output_file = ""
        self.sheet_names = None
        self.color_column = "B"
        self.hierarchy_column = "A"
        self.min_row = 11
        self.scan_columns_by_row = None
        self.font = {'name': 'Times New Roman', 'size': 14, 'bold': False, 'italic': False, 'underline': False}
        self.border_style = 'thin'
        self.bold_levels = [1, 2]
        self.column_formats = {
            'E': '#,##0.00',
            'F:I': '#,##0'
        }
        self.wrap_text_columns = ['B']
        self.alignment_rules = [
            ('A:B', 'center', 'left')
        ]
        self.fill_color = None
        self.text_color = None
        self.grid_enabled = False
        self.stages = {
            'grouping': True,
            'hierarchy': True,
            'hierarchy_colors': True,
            'wrap_text': False,
            'alignment': False,
            'formatting': False,
            'number_formats': False,
            'large_file_mode': False,
        }

    def to_dict(self):
        return {
            "input_file": self.input_file,
            "output_file": self.output_file,
            "sheet_names": self.sheet_names,
            "color_column": self.color_column,
            "hierarchy_column": self.hierarchy_column,
            "min_row": self.min_row,
            "scan_columns_by_row": self.scan_columns_by_row,
            "font": self.font,
            "border_style": self.border_style,
            "bold_levels": self.bold_levels,
            "column_formats": self.column_formats,
            "wrap_text_columns": self.wrap_text_columns,
            "alignment_rules": self.alignment_rules,
            "fill_color": self.fill_color,
            "text_color": self.text_color,
            "grid_enabled": self.grid_enabled,
            "stages": self.stages
        }

    def from_dict(self, data):
        self.input_file = data.get("input_file", "")
        self.output_file = data.get("output_file", "")
        self.sheet_names = data.get("sheet_names", None)
        self.color_column = data.get("color_column", "B")
        self.hierarchy_column = data.get("hierarchy_column", "A")
        self.min_row = data.get("min_row", 11)
        self.scan_columns_by_row = data.get("scan_columns_by_row", None)
        self.font = data.get("font", {'name': 'Times New Roman', 'size': 14})
        self.border_style = data.get("border_style", "thin")
        self.bold_levels = data.get("bold_levels", [1, 2])
        self.column_formats = data.get("column_formats", {})
        self.wrap_text_columns = data.get("wrap_text_columns", [])
        self.alignment_rules = data.get("alignment_rules", [])
        self.fill_color = data.get("fill_color", None)
        self.text_color = data.get("text_color", None)
        self.grid_enabled = data.get("grid_enabled", False)
        self.stages = data.get("stages", {})


# ======================
# ФОРМАТ ПАНЕЛЬ (БЕЗ ЦВЕТА ФОНА И ТЕКСТА)
# ======================

class FormatPanel(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent = parent
        self.layout = QGridLayout()
        self.setLayout(self.layout)

        row = 0

        self.layout.addWidget(QLabel("Шрифт:"), row, 0)
        self.font_combo = QFontComboBox()
        self.font_combo.currentFontChanged.connect(self.on_font_changed)
        self.layout.addWidget(self.font_combo, row, 1)

        self.layout.addWidget(QLabel("Размер:"), row, 2)
        self.size_spin = QSpinBox()
        self.size_spin.setMinimum(6)
        self.size_spin.setMaximum(72)
        self.size_spin.setValue(14)
        self.size_spin.valueChanged.connect(self.on_size_changed)
        self.layout.addWidget(self.size_spin, row, 3)

        row += 1

        self.bold_btn = QToolButton()
        self.bold_btn.setText("Ж")
        self.bold_btn.setCheckable(True)
        self.bold_btn.clicked.connect(self.toggle_bold)
        self.layout.addWidget(self.bold_btn, row, 0)

        self.italic_btn = QToolButton()
        self.italic_btn.setText("К")
        self.italic_btn.setCheckable(True)
        self.italic_btn.clicked.connect(self.toggle_italic)
        self.layout.addWidget(self.italic_btn, row, 1)

        self.underline_btn = QToolButton()
        self.underline_btn.setText("У")
        self.underline_btn.setCheckable(True)
        self.underline_btn.clicked.connect(self.toggle_underline)
        self.layout.addWidget(self.underline_btn, row, 2)

        row += 1

        self.layout.addWidget(QLabel("Жирные уровни (через запятую):"), row, 0, 1, 2)
        self.bold_levels_edit = QLineEdit("1,2")
        self.bold_levels_edit.setToolTip("Укажите уровни иерархии, которые будут жирными. Например: 1,2,3")
        self.bold_levels_edit.textChanged.connect(self.on_bold_levels_changed)
        self.layout.addWidget(self.bold_levels_edit, row, 2, 1, 2)

        row += 1

        self.layout.addWidget(QLabel("Выравнивание:"), row, 0)
        self.align_combo = QComboBox()
        self.align_combo.addItems(["Left", "Center", "Right", "Justify"])
        self.layout.addWidget(self.align_combo, row, 1)

        self.layout.addWidget(QLabel("Границы:"), row, 2)
        border_styles = ["thin", "medium", "thick", "double", "hair", "dashed", "dotted", "none"]
        self.border_combo = QComboBox()
        self.border_combo.addItems(border_styles)
        self.border_combo.setCurrentText("thin")
        self.layout.addWidget(self.border_combo, row, 3)

        row += 1

        # ✅ УДАЛЕНО: Цвет фона и Цвет текста

        self.preview_label = PreviewLabel()
        self.layout.addWidget(self.preview_label, row, 0, 1, 4)  # Занимает всю ширину

        self.layout.setRowStretch(row+1, 1)
        self.update_preview()

    def on_font_changed(self, font):
        self.parent.config.font['name'] = font.family()
        self.update_preview()

    def on_size_changed(self, value):
        self.parent.config.font['size'] = value
        self.update_preview()

    def toggle_bold(self):
        self.parent.config.font['bold'] = self.bold_btn.isChecked()
        self.update_preview()

    def toggle_italic(self):
        self.parent.config.font['italic'] = self.italic_btn.isChecked()
        self.update_preview()

    def toggle_underline(self):
        self.parent.config.font['underline'] = self.underline_btn.isChecked()
        self.update_preview()

    def on_bold_levels_changed(self, text):
        try:
            levels = [int(x.strip()) for x in text.split(",") if x.strip().isdigit()]
            self.parent.config.bold_levels = levels
            self.bold_levels_edit.setStyleSheet("")
        except:
            self.bold_levels_edit.setStyleSheet("background-color: #ffdddd;")

    def update_preview(self, font_name=None, font_size=None, bold=None, italic=None, underline=None, text_color=None, bg_color=None):
        if font_name is None: font_name = self.parent.config.font.get('name', 'Times New Roman')
        if font_size is None: font_size = self.parent.config.font.get('size', 14)
        if bold is None: bold = self.parent.config.font.get('bold', False)
        if italic is None: italic = self.parent.config.font.get('italic', False)
        if underline is None: underline = self.parent.config.font.get('underline', False)
        # ✅ Цвета больше не используются в превью
        self.preview_label.update_preview(font_name, font_size, bold, italic, underline, None, None)


# ======================
# РЕДАКТОР ЧИСЛОВЫХ ФОРМАТОВ
# ======================

class ColumnFormatEditor(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent = parent
        layout = QVBoxLayout(self)

        self.table = QTableWidget(0, 3)
        self.table.setHorizontalHeaderLabels(["Диапазон столбцов", "Формат", "Предпросмотр"])
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        layout.addWidget(self.table)

        btn_layout = QHBoxLayout()
        add_btn = QPushButton("Добавить")
        add_btn.clicked.connect(self.add_row)
        remove_btn = QPushButton("Удалить")
        remove_btn.clicked.connect(self.remove_row)
        btn_layout.addWidget(add_btn)
        btn_layout.addWidget(remove_btn)
        layout.addLayout(btn_layout)

        self.table.cellChanged.connect(self.on_cell_changed)

    def load_data(self, data):
        self.table.cellChanged.disconnect()
        self.table.setRowCount(0)
        for col_range, fmt in data.items():
            row = self.table.rowCount()
            self.table.insertRow(row)
            self.table.setItem(row, 0, QTableWidgetItem(col_range))
            self.table.setItem(row, 1, QTableWidgetItem(fmt))
            preview = NumberFormatPreview()
            preview.update_preview(fmt)
            self.table.setCellWidget(row, 2, preview)
        self.table.cellChanged.connect(self.on_cell_changed)

    def save_data(self):
        result = {}
        for row in range(self.table.rowCount()):
            col_range = self.table.item(row, 0)
            fmt = self.table.item(row, 1)
            if col_range and fmt:
                result[col_range.text()] = fmt.text()
        return result

    def add_row(self):
        row = self.table.rowCount()
        self.table.insertRow(row)
        self.table.setItem(row, 0, QTableWidgetItem("A"))
        self.table.setItem(row, 1, QTableWidgetItem("#,##0.00"))
        preview = NumberFormatPreview()
        preview.update_preview("#,##0.00")
        self.table.setCellWidget(row, 2, preview)

    def remove_row(self):
        row = self.table.currentRow()
        if row >= 0:
            self.table.removeRow(row)

    def on_cell_changed(self, row, column):
        if column == 1:
            fmt_item = self.table.item(row, 1)
            if fmt_item:
                fmt = fmt_item.text()
                preview_widget = self.table.cellWidget(row, 2)
                if not preview_widget:
                    preview_widget = NumberFormatPreview()
                    self.table.setCellWidget(row, 2, preview_widget)
                preview_widget.update_preview(fmt)


# ======================
# РЕДАКТОР ВЫРАВНИВАНИЯ
# ======================

class AlignmentEditor(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent = parent
        layout = QVBoxLayout(self)

        self.table = QTableWidget(0, 3)
        self.table.setHorizontalHeaderLabels(["Диапазон", "Вертикально", "Горизонтально"])
        self.table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(self.table)

        v_aligns = ["top", "center", "bottom", "justify"]
        h_aligns = ["left", "center", "right", "justify"]

        btn_layout = QHBoxLayout()
        add_btn = QPushButton("Добавить")
        add_btn.clicked.connect(lambda: self.add_row(h_aligns[0], v_aligns[0]))
        remove_btn = QPushButton("Удалить")
        remove_btn.clicked.connect(self.remove_row)
        btn_layout.addWidget(add_btn)
        btn_layout.addWidget(remove_btn)
        layout.addLayout(btn_layout)

        self.v_aligns = v_aligns
        self.h_aligns = h_aligns

    def load_data(self, data):
        self.table.setRowCount(0)
        for rule in data:
            if len(rule) != 3:
                continue
            col_range, v, h = rule
            row = self.table.rowCount()
            self.table.insertRow(row)
            self.table.setItem(row, 0, QTableWidgetItem(col_range))
            combo_v = QComboBox()
            combo_v.addItems(self.v_aligns)
            combo_v.setCurrentText(v)
            self.table.setCellWidget(row, 1, combo_v)
            combo_h = QComboBox()
            combo_h.addItems(self.h_aligns)
            combo_h.setCurrentText(h)
            self.table.setCellWidget(row, 2, combo_h)

    def save_data(self):
        result = []
        for row in range(self.table.rowCount()):
            col_range = self.table.item(row, 0)
            combo_v = self.table.cellWidget(row, 1)
            combo_h = self.table.cellWidget(row, 2)
            if col_range and combo_v and combo_h:
                result.append([col_range.text(), combo_v.currentText(), combo_h.currentText()])
        return result

    def add_row(self, v_default, h_default):
        row = self.table.rowCount()
        self.table.insertRow(row)
        self.table.setItem(row, 0, QTableWidgetItem("A:B"))
        combo_v = QComboBox()
        combo_v.addItems(self.v_aligns)
        combo_v.setCurrentText(v_default)
        self.table.setCellWidget(row, 1, combo_v)
        combo_h = QComboBox()
        combo_h.addItems(self.h_aligns)
        combo_h.setCurrentText(h_default)
        self.table.setCellWidget(row, 2, combo_h)

    def remove_row(self):
        row = self.table.currentRow()
        if row >= 0:
            self.table.removeRow(row)


# ======================
# ПОТОК ОБРАБОТКИ
# ======================

class WorkerThread(QThread):
    log_signal = pyqtSignal(str)
    finished_signal = pyqtSignal(bool, str)
    stopped = False

    def __init__(self, config, sheet_names):
        super().__init__()
        self.config = config
        self.sheet_names = sheet_names

    def run(self):
        if not self.sheet_names:
            self.finished_signal.emit(False, "Нет выбранных листов")
            return

        total = len(self.sheet_names)
        for i, sheet_name in enumerate(self.sheet_names):
            if self.stopped:
                self.log_signal.emit("🛑 Обработка остановлена пользователем.")
                self.finished_signal.emit(False, "Остановлено пользователем")
                return

            self.log_signal.emit(f"📋 Обработка листа {i+1}/{total}: {sheet_name}")

            temp_config = Config()
            temp_config.__dict__.update(self.config.__dict__)
            temp_config.sheet_names = [sheet_name]

            try:
                success, message = process_excel(temp_config.__dict__, self.log_signal.emit)
                if not success:
                    self.finished_signal.emit(False, f"Ошибка на листе {sheet_name}: {message}")
                    return
            except Exception as e:
                self.finished_signal.emit(False, f"Исключение на листе {sheet_name}: {str(e)}")
                return

        self.finished_signal.emit(True, "Обработка завершена успешно.")

    def stop(self):
        self.stopped = True


# ======================
# ВКЛАДКА КОНФИГА
# ======================

class ConfigTab(QWidget):
    def __init__(self, parent=None, config_name="Конфиг 1"):
        super().__init__(parent)
        self.parent = parent
        self.config = Config()
        self.config_name = config_name
        self.worker = None
        self.last_input_file = None
        self.initUI()

    def initUI(self):
        main_layout = QVBoxLayout(self)

        # Этапы обработки — 3 колонки + "Большой файл" с подсказкой
        stages_group = QGroupBox("Этапы обработки")
        stages_layout = QGridLayout()

        self.stage_checks = {}
        stages = [
            ("Группировка", 'grouping'),
            ("Иерархия", 'hierarchy'),
            ("Цвет в иерархии", 'hierarchy_colors'),
            ("Перенос текста", 'wrap_text'),
            ("Выравнивание", 'alignment'),
            ("Форматирование", 'formatting'),
            ("Числовые форматы", 'number_formats'),
            ("Большой файл", 'large_file_mode'),
        ]

        for i, (label, key) in enumerate(stages):
            check = QCheckBox(label)
            check.setChecked(key in ['grouping', 'hierarchy'])
            check.stateChanged.connect(self.toggle_sections)
            self.stage_checks[key] = check
            stages_layout.addWidget(check, i // 3, i % 3)
            # ✅ Добавляем подсказку к "Большой файл"
            if key == 'large_file_mode':
                check.setToolTip("Включает оптимизацию для больших файлов.\nАвтоматически обрабатывает все столбцы до цветового включительно.")

        stages_group.setLayout(stages_layout)
        main_layout.addWidget(stages_group)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll_content = QWidget()
        scroll_content.setStyleSheet("background-color: #2b2b2b;")
        scroll_layout = QVBoxLayout(scroll_content)
        scroll.setWidget(scroll_content)
        main_layout.addWidget(scroll)

        # Format Panel
        self.format_panel = FormatPanel(self)
        self.format_panel_group = QGroupBox("Форматирование шрифта")
        fp_layout = QVBoxLayout()
        fp_layout.addWidget(self.format_panel)
        self.format_panel_group.setLayout(fp_layout)
        scroll_layout.addWidget(self.format_panel_group)

        # Input File
        input_layout = QHBoxLayout()
        self.input_line = QLineEdit()
        self.input_line.setPlaceholderText("Выберите входной Excel файл...")
        input_btn = QPushButton("Выбрать файл...")
        input_btn.clicked.connect(self.select_input_file)
        input_btn.setStyleSheet("""
            QPushButton {
                background-color: #4caf50;
                color: white;
                padding: 8px;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        input_layout.addWidget(QLabel("Входной файл:"))
        input_layout.addWidget(self.input_line)
        input_layout.addWidget(input_btn)
        input_group = QGroupBox("Файл")
        input_group.setLayout(input_layout)
        scroll_layout.addWidget(input_group)

        # Output File
        output_layout = QHBoxLayout()
        self.output_line = QLineEdit()
        self.output_line.setPlaceholderText("Автоматически: <имя>_обработанный.xlsx")
        output_btn = QPushButton("Выбрать место...")
        output_btn.clicked.connect(self.select_output_file)
        output_layout.addWidget(QLabel("Выходной файл:"))
        output_layout.addWidget(self.output_line)
        output_layout.addWidget(output_btn)
        output_group = QGroupBox("Сохранение")
        output_group.setLayout(output_layout)
        scroll_layout.addWidget(output_group)

        # Parameters Group
        params_group = QGroupBox("Параметры обработки")
        params_layout = QGridLayout()
        params_layout.addWidget(QLabel("Цветовой столбец:"), 0, 0)
        self.color_col_edit = QLineEdit("B")
        params_layout.addWidget(self.color_col_edit, 0, 1)
        params_layout.addWidget(QLabel("Столбец иерархии:"), 1, 0)
        self.hierarchy_col_edit = QLineEdit("A")
        params_layout.addWidget(self.hierarchy_col_edit, 1, 1)
        params_layout.addWidget(QLabel("Начальная строка:"), 2, 0)
        self.min_row_spin = QSpinBox()
        self.min_row_spin.setMinimum(1)
        self.min_row_spin.setMaximum(10000)
        self.min_row_spin.setValue(11)
        params_layout.addWidget(self.min_row_spin, 2, 1)
        params_group.setLayout(params_layout)
        scroll_layout.addWidget(params_group)

        # ✅ Блок "Сканировать столбцы по строке" — УДАЛЁН

        # Выбор листов и логи — на одной высоте
        sheets_logs_layout = QHBoxLayout()

        sheet_group = QGroupBox("Выбор листов")
        self.sheet_list = QListWidget()
        self.sheet_list.setSelectionMode(QListWidget.MultiSelection)
        sheet_font = self.sheet_list.font()
        sheet_font.setPointSize(sheet_font.pointSize() + 2)
        self.sheet_list.setFont(sheet_font)
        sheet_group_layout = QVBoxLayout()
        sheet_group_layout.addWidget(self.sheet_list)
        sheet_group.setLayout(sheet_group_layout)
        sheets_logs_layout.addWidget(sheet_group, 1)

        log_group = QGroupBox("Логи")
        log_layout = QVBoxLayout()
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setPlaceholderText("Логи будут отображаться здесь...")
        log_layout.addWidget(self.log_text)
        log_group.setLayout(log_layout)
        log_group.setMinimumHeight(250)
        sheets_logs_layout.addWidget(log_group, 2)

        scroll_layout.addLayout(sheets_logs_layout)

        # Editors
        self.editors_group = QGroupBox("Редакторы форматов и выравнивания")
        editors_layout = QHBoxLayout()
        self.column_format_editor = ColumnFormatEditor(self)
        self.alignment_editor = AlignmentEditor(self)
        editors_layout.addWidget(self.column_format_editor)
        editors_layout.addWidget(self.alignment_editor)
        self.editors_group.setLayout(editors_layout)
        scroll_layout.addWidget(self.editors_group)

        # Start/Stop Button
        self.start_stop_btn = QPushButton("▶️ Запустить обработку")
        self.start_stop_btn.clicked.connect(self.toggle_start_stop)
        self.start_stop_btn.setStyleSheet("""
            QPushButton {
                background-color: #4caf50;
                color: white;
                font-size: 16px;
                padding: 12px;
                border-radius: 6px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        scroll_layout.addWidget(self.start_stop_btn)

        # Секции для скрытия
        self.section_widgets = {
            'hierarchy_colors': [self.format_panel_group],
            'wrap_text': [self.format_panel_group],
            'alignment': [self.format_panel_group, self.editors_group],
            'formatting': [self.format_panel_group, self.editors_group],
            'number_formats': [self.editors_group],
            # ✅ 'large_file_mode' не показывает блоков
        }

        self.toggle_sections()
        self.input_line.textChanged.connect(self.on_input_file_changed)

    def toggle_sections(self):
        visible_widgets = set()
        for key, widgets in self.section_widgets.items():
            if self.stage_checks[key].isChecked():
                for w in widgets:
                    visible_widgets.add(w)
        all_widgets = set(w for widgets in self.section_widgets.values() for w in widgets)
        for widget in all_widgets:
            widget.setVisible(widget in visible_widgets)

    def on_input_file_changed(self, new_file):
        if new_file and new_file != self.last_input_file:
            self.last_input_file = new_file
            if not self.output_line.text() or (self.output_line.text().endswith("_обработанный.xlsx") and self.last_input_file):
                p = os.path.splitext(new_file)[0]
                self.output_line.setText(p + "_обработанный.xlsx")

    def select_input_file(self):
        file, _ = QFileDialog.getOpenFileName(self, "Выберите Excel файл", "", "Excel Files (*.xlsx *.xls)")
        if file:
            self.input_line.setText(file)
            p = os.path.splitext(file)[0]
            self.output_line.setText(p + "_обработанный.xlsx")

            size = os.path.getsize(file)
            mb = size / (1024 * 1024)
            if mb < 10:
                est = "менее 5 секунд"
            elif mb < 50:
                est = "10-30 секунд"
            elif mb < 200:
                est = "30-90 секунд"
            else:
                est = "более 2 минут"

            self.log(f"⏳ Загружается файл ({mb:.1f} МБ). Ориентировочное время: {est}...")
            QApplication.processEvents()

            self.load_sheets()

    def select_output_file(self):
        file, _ = QFileDialog.getSaveFileName(self, "Сохранить как", "", "Excel Files (*.xlsx)")
        if file:
            if not file.endswith(".xlsx"):
                file += ".xlsx"
            self.output_line.setText(file)

    def load_sheets(self):
        file = self.input_line.text()
        if not file or not os.path.exists(file):
            return
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file, read_only=True)
            self.sheet_list.clear()
            for name in wb.sheetnames:
                item = QListWidgetItem(name)
                item.setCheckState(Qt.Checked)
                self.sheet_list.addItem(item)
            wb.close()
            self.log("✅ Файл успешно загружен. Листы готовы к обработке.")
        except Exception as e:
            self.log(f"❌ Ошибка загрузки листов: {e}")

    def toggle_start_stop(self):
        if self.start_stop_btn.text() == "▶️ Запустить обработку":
            self.start_processing()
        else:
            self.stop_processing()

    def start_processing(self):
        if not self.input_line.text():
            self.log("❌ Пожалуйста, выберите входной файл.")
            return

        output_file = self.output_line.text() or (os.path.splitext(self.input_line.text())[0] + "_обработанный.xlsx")
        if os.path.exists(output_file):
            try:
                with open(output_file, 'a'):
                    pass
            except PermissionError:
                QMessageBox.warning(
                    self,
                    "Файл занят",
                    "❌ Файл уже открыт в Excel или другом приложении.\n\n"
                    "Пожалуйста, закройте его и попробуйте снова."
                )
                return

        selected_sheets = []
        for i in range(self.sheet_list.count()):
            item = self.sheet_list.item(i)
            if item.checkState() == Qt.Checked:
                selected_sheets.append(item.text())
        if not selected_sheets:
            self.log("❌ Нет выбранных листов.")
            return

        self.config.input_file = self.input_line.text()
        self.config.output_file = output_file
        self.config.color_column = self.color_col_edit.text().strip().upper()
        self.config.hierarchy_column = self.hierarchy_col_edit.text().strip().upper()
        self.config.min_row = self.min_row_spin.value()

        # ✅ Упрощённая логика: просто флаг
        self.config.scan_columns_by_row = 1 if self.stage_checks['large_file_mode'].isChecked() else None

        for key, check in self.stage_checks.items():
            self.config.stages[key] = check.isChecked()

        self.config.column_formats = self.column_format_editor.save_data()
        self.config.alignment_rules = self.alignment_editor.save_data()

        self.start_stop_btn.setText("⏹️ Стоп")
        self.start_stop_btn.setStyleSheet("""
            QPushButton {
                background-color: #d32f2f;
                color: white;
                font-size: 16px;
                padding: 12px;
                border-radius: 6px;
            }
            QPushButton:hover {
                background-color: #c62828;
            }
        """)
        self.log_text.clear()
        self.log("🚀 Начинаем обработку...")

        self.worker = WorkerThread(self.config, selected_sheets)
        self.worker.log_signal.connect(self.log)
        self.worker.finished_signal.connect(self.on_finished)
        self.worker.start()

    def stop_processing(self):
        if self.worker:
            self.worker.stop()
            self.start_stop_btn.setEnabled(False)
            self.log("🛑 Запрос на остановку отправлен...")

    def on_finished(self, success, message):
        self.start_stop_btn.setText("▶️ Запустить обработку")
        self.start_stop_btn.setStyleSheet("""
            QPushButton {
                background-color: #4caf50;
                color: white;
                font-size: 16px;
                padding: 12px;
                border-radius: 6px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        self.start_stop_btn.setEnabled(True)

        if success:
            self.log("🎉 Обработка завершена успешно!")
            try:
                for proc in psutil.process_iter(['pid', 'name']):
                    if proc.info['name'] in ['EXCEL.EXE', 'excel.exe']:
                        proc.terminate()
                        self.log(f"✅ Процесс Excel (PID {proc.info['pid']}) завершён.")
            except Exception as e:
                self.log(f"⚠️ Не удалось завершить процессы Excel: {e}")
        else:
            if message != "Остановлено пользователем":
                self.log(f"❌ Ошибка: {message}")

    def log(self, message):
        self.log_text.append(message)
        self.log_text.verticalScrollBar().setValue(self.log_text.verticalScrollBar().maximum())


# ======================
# ГЛАВНОЕ ОКНО
# ======================

class ExcelProcessorGUI(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Chik-chik")
        self.setGeometry(100, 100, 1200, 1100)

        self.setWindowFlags(Qt.FramelessWindowHint)

        app = QApplication.instance()
        font = app.font()
        font.setPointSize(font.pointSize() + 3)
        app.setFont(font)

        app.setStyleSheet("""
            QMainWindow, QWidget {
                background-color: #2b2b2b;
                color: white;
            }
            QGroupBox {
                color: white;
                font-weight: bold;
                border: 1px solid #444;
                border-radius: 6px;
                margin-top: 10px;
                padding: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
                color: white;
            }
            QLabel {
                color: white;
            }
            QLineEdit, QComboBox, QSpinBox, QFontComboBox {
                background-color: #3a3a3a;
                color: white;
                border: 1px solid #666;
                padding: 5px;
                border-radius: 4px;
            }
            QCheckBox, QRadioButton {
                color: white;
                background-color: transparent;
            }
            QCheckBox::indicator {
                width: 18px;
                height: 18px;
                border: 2px solid #888;
                border-radius: 4px;
            }
            QCheckBox::indicator:checked {
                background-color: #5a5a5a;
                border: 2px solid #aaa;
            }
            QPushButton {
                background-color: #4a4a4a;
                color: white;
                border: 1px solid #666;
                padding: 8px;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #5a5a5a;
            }
            QTextEdit {
                background-color: #1e1e1e;
                color: white;
                border: 1px solid #555;
                border-radius: 4px;
            }
            QListWidget {
                background-color: #3a3a3a;
                color: white;
                border: 1px solid #555;
                border-radius: 4px;
            }
            QTableWidget {
                background-color: #3a3a3a;
                color: white;
                border: 1px solid #555;
                gridline-color: #666;
            }
            QHeaderView::section {
                background-color: #4a4a4a;
                color: white;
                padding: 4px;
                border: 1px solid #444;
            }
            QScrollBar:vertical {
                background: #2b2b2b;
                width: 12px;
                margin: 0;
            }
            QScrollBar::handle:vertical {
                background: #666;
                min-height: 20px;
                border-radius: 6px;
            }
            QScrollBar::handle:vertical:hover {
                background: #888;
            }
            QSpinBox::up-button, QSpinBox::down-button {
                width: 20px;
                background-color: #4a4a4a;
                border: 1px solid #666;
            }
            QSpinBox::up-button:hover, QSpinBox::down-button:hover {
                background-color: #5a5a5a;
            }
        """)

        central_widget = QWidget()
        central_layout = QVBoxLayout(central_widget)
        central_layout.setContentsMargins(0, 0, 0, 0)
        central_layout.setSpacing(0)

        self.title_bar = QWidget()
        self.title_bar.setFixedHeight(40)
        self.title_bar.setStyleSheet("background-color: #1a1a1a; border-top-left-radius: 8px; border-top-right-radius: 8px; border-bottom: 2px solid #444;")

        title_layout = QHBoxLayout(self.title_bar)
        title_layout.setContentsMargins(15, 0, 15, 0)

        title_label = QLabel("Chik-chik")
        title_label.setStyleSheet("color: white; font-weight: bold; font-size: 14px;")
        title_layout.addWidget(title_label)

        title_layout.addStretch()

        minimize_btn = QToolButton()
        minimize_btn.setText("—")
        minimize_btn.clicked.connect(self.showMinimized)
        minimize_btn.setStyleSheet("""
            QToolButton {
                color: white;
                background: transparent;
                font-size: 18px;
                border: none;
                padding: 5px;
                border-radius: 4px;
            }
            QToolButton:hover {
                background-color: #333;
            }
        """)

        close_btn = QToolButton()
        close_btn.setText("✕")
        close_btn.clicked.connect(self.close)
        close_btn.setStyleSheet("""
            QToolButton {
                color: white;
                background: transparent;
                font-size: 18px;
                border: none;
                padding: 5px;
                border-radius: 4px;
            }
            QToolButton:hover {
                background-color: #aa3333;
            }
        """)

        title_layout.addWidget(minimize_btn)
        title_layout.addWidget(close_btn)

        central_layout.addWidget(self.title_bar)

        content = QWidget()
        content.setStyleSheet("background-color: #2b2b2b; border-bottom-left-radius: 8px; border-bottom-right-radius: 8px;")
        content_layout = QVBoxLayout(content)
        content_layout.setContentsMargins(15, 15, 15, 15)
        content_layout.setSpacing(15)

        self.tabs = QTabWidget()
        self.tabs.setTabsClosable(True)
        self.tabs.tabCloseRequested.connect(self.close_tab)
        self.tabs.setStyleSheet("""
            QTabWidget::pane {
                border: 1px solid #444;
                border-radius: 6px;
                background: #2b2b2b;
            }
            QTabBar::tab {
                background: #3a3a3a;
                color: white;
                padding: 8px 25px;
                margin: 2px;
                border-top-left-radius: 6px;
                border-top-right-radius: 6px;
                min-width: 120px;
            }
            QTabBar::tab:selected {
                background: #4a4a4a;
                font-weight: bold;
            }
            QTabBar::tab:hover {
                background: #454545;
            }
        """)
        content_layout.addWidget(self.tabs)

        btn_layout = QHBoxLayout()
        add_tab_btn = QPushButton("➕ Добавить конфиг")
        add_tab_btn.clicked.connect(self.add_tab)
        save_btn = QPushButton("💾 Сохранить настройки")
        save_btn.clicked.connect(self.save_settings)
        load_btn = QPushButton("📂 Загрузить настройки")
        load_btn.clicked.connect(self.load_settings)
        exit_btn = QPushButton("🚪 Выход")
        exit_btn.clicked.connect(self.confirm_close)
        exit_btn.setStyleSheet("""
            QPushButton {
                background-color: #aa3333;
                color: white;
                font-size: 16px;
                padding: 10px;
                border-radius: 6px;
            }
            QPushButton:hover {
                background-color: #cc4444;
            }
        """)

        btn_layout.addWidget(add_tab_btn)
        btn_layout.addWidget(save_btn)
        btn_layout.addWidget(load_btn)
        btn_layout.addWidget(exit_btn)
        content_layout.addLayout(btn_layout)

        central_layout.addWidget(content)
        self.setCentralWidget(central_widget)

        self.title_bar.mousePressEvent = self.mousePressEvent
        self.title_bar.mouseMoveEvent = self.mouseMoveEvent
        self.dragPos = None

        self.add_tab()

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton and event.y() < self.title_bar.height():
            self.dragPos = event.globalPos()

    def mouseMoveEvent(self, event):
        if self.dragPos and event.buttons() == Qt.LeftButton:
            self.move(self.pos() + event.globalPos() - self.dragPos)
            self.dragPos = event.globalPos()

    def add_tab(self):
        index = self.tabs.count() + 1
        tab = ConfigTab(self, f"Конфиг {index}")
        self.tabs.addTab(tab, tab.config_name)
        self.tabs.setCurrentWidget(tab)

    def close_tab(self, index):
        if self.tabs.count() > 1:
            self.tabs.removeTab(index)
        else:
            QMessageBox.warning(self, "Предупреждение", "Нельзя закрыть последнюю вкладку!")

    def save_settings(self):
        configs = []
        for i in range(self.tabs.count()):
            tab = self.tabs.widget(i)
            config_data = tab.config.to_dict()
            config_data["__tab_name__"] = self.tabs.tabText(i)
            configs.append(config_data)

        file, _ = QFileDialog.getSaveFileName(self, "Сохранить настройки", "", "JSON Files (*.json)")
        if file:
            if not file.endswith(".json"):
                file += ".json"
            try:
                with open(file, "w", encoding="utf-8") as f:
                    json.dump(configs, f, indent=4, ensure_ascii=False)
                QMessageBox.information(self, "Успех", "Настройки сохранены!")
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить: {str(e)}")

    def load_settings(self):
        file, _ = QFileDialog.getOpenFileName(self, "Загрузить настройки", "", "JSON Files (*.json)")
        if not file:
            return

        try:
            with open(file, "r", encoding="utf-8") as f:
                configs = json.load(f)

            while self.tabs.count() > 1:
                self.tabs.removeTab(1)

            for i, config_data in enumerate(configs):
                if i == 0:
                    tab = self.tabs.widget(0)
                else:
                    tab = ConfigTab(self)
                    self.tabs.addTab(tab, "Загружено...")
                tab.config.from_dict(config_data)
                tab_name = config_data.get("__tab_name__", f"Конфиг {i+1}")
                self.tabs.setTabText(self.tabs.indexOf(tab), tab_name)

                tab.input_line.setText(tab.config.input_file)
                tab.output_line.setText(tab.config.output_file or "")
                tab.color_col_edit.setText(tab.config.color_column)
                tab.hierarchy_col_edit.setText(tab.config.hierarchy_column)
                tab.min_row_spin.setValue(tab.config.min_row)

                for key, check in tab.stage_checks.items():
                    check.setChecked(tab.config.stages.get(key, False))

                # ✅ Не восстанавливаем scan_row_spin — он больше не используется

                tab.column_format_editor.load_data(tab.config.column_formats)
                tab.alignment_editor.load_data(tab.config.alignment_rules)

                if hasattr(tab.config, 'font'):
                    font = tab.config.font
                    tab.format_panel.font_combo.setCurrentFont(QFont(font.get('name', 'Times New Roman')))
                    tab.format_panel.size_spin.setValue(font.get('size', 14))
                    tab.format_panel.bold_btn.setChecked(font.get('bold', False))
                    tab.format_panel.italic_btn.setChecked(font.get('italic', False))
                    tab.format_panel.underline_btn.setChecked(font.get('underline', False))

                # ✅ Цвета больше не восстанавливаются — их нет в интерфейсе

                tab.format_panel.border_combo.setCurrentText(tab.config.border_style)

                if hasattr(tab.config, 'bold_levels'):
                    levels_str = ",".join(map(str, tab.config.bold_levels))
                    tab.format_panel.bold_levels_edit.setText(levels_str)

            QMessageBox.information(self, "Успех", "Настройки загружены!")

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось загрузить: {str(e)}")

    def confirm_close(self):
        reply = QMessageBox.question(
            self,
            "Подтверждение выхода",
            "Вы уверены, что хотите выйти из Chik-chik?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            self.close()

    def closeEvent(self, event):
        reply = QMessageBox.question(
            self,
            "Подтверждение выхода",
            "Вы уверены, что хотите выйти из Chik-chik?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            settings = QSettings("Chik-chik", "Geometry")
            settings.setValue("geometry", self.saveGeometry())
            event.accept()
        else:
            event.ignore()


# ======================
# ЗАПУСК
# ======================

if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyle(QStyleFactory.create("Fusion"))
    window = ExcelProcessorGUI()
    window.show()
    sys.exit(app.exec_())